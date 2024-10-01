import openpyxl
from openpyxl.styles import Font

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Salary Prediction"

# Add headers and descriptions
ws['A1'] = "Enter Gross Annual Income in B2, Bonus in B3, Spot Cash Awards in B4, and Increment Percentage in B5."
ws['A2'] = "Gross Annual Income"
ws['B2'] = 0  # Placeholder for user input

ws['A3'] = "Bonus (Credited in March)"
ws['B3'] = 0  # Placeholder for user input

ws['A4'] = "Spot Cash Award"
ws['B4'] = 0  # Placeholder for user input

ws['A5'] = "Increment Percentage"
ws['B5'] = 0  # Placeholder for user input

ws['A6'] = "New Gross Annual Income"
ws['B6'] = "=$B$2*(1+$B$5/100)"  # Formula to calculate new salary after increment

ws['A7'] = "Number of Unused Leaves"
ws['B7'] = 0  # Placeholder for user input

ws['A8'] = "Per-Day Salary"
ws['B8'] = "=$B$2/365"  # Formula to calculate per-day salary from gross annual income

# Add month labels
months = ["March", "April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February"]
headers = ["Month", "Basic Salary", "PF Deduction (12%)", "Tax Deduction", "Variable Bonus", "Leave Encashment", "Spot Cash Award", "Total Take-Home"]

# Add headers to the table
for col_num, header in enumerate(headers, 1):
    ws.cell(row=9, column=col_num).value = header
    ws.cell(row=9, column=col_num).font = Font(bold=True)

# Add months to the rows
for i, month in enumerate(months, start=10):
    ws.cell(row=i, column=1).value = month

# Add formulas for each month for current year
for row in range(10, 22):
    # Basic Salary
    ws.cell(row=row, column=2).value = "=$B$2/12"
    
    # PF Deduction (12% of Basic Salary)
    ws.cell(row=row, column=3).value = "=$B$2/12*0.12"
    
    # Tax Deduction (Assume 10% tax as a placeholder, you can replace with actual tax slab formulas)
    ws.cell(row=row, column=4).value = "=$B$2*0.10/12"
    
    # Variable Bonus in March only
    if row == 10:  # March
        ws.cell(row=row, column=5).value = "=$B$3"
    else:
        ws.cell(row=row, column=5).value = 0
    
    # Leave Encashment in January
    if row == 19:  # January
        ws.cell(row=row, column=6).value = "=$B$7*$B$8"
    else:
        ws.cell(row=row, column=6).value = 0
    
    # Spot Cash Award (can be manually entered in respective months)
    ws.cell(row=row, column=7).value = "=IF(ISBLANK($B$4),0,$B$4)"
    
    # Total Take-Home
    ws.cell(row=row, column=8).value = "=B{row}-C{row}-D{row}+E{row}+F{row}+G{row}".format(row=row)

# Add section for next year prediction with increment
ws['A23'] = "Next Year Salary Prediction (With Increment)"
for col_num, header in enumerate(headers, 1):
    ws.cell(row=24, column=col_num).value = header
    ws.cell(row=24, column=col_num).font = Font(bold=True)

# Add months for next year's projection
for i, month in enumerate(months, start=25):
    ws.cell(row=i, column=1).value = month

# Add formulas for next year based on incremented salary
for row in range(25, 37):
    # Basic Salary (New Gross Annual Income / 12)
    ws.cell(row=row, column=2).value = "=$B$6/12"
    
    # PF Deduction (12% of new Basic Salary)
    ws.cell(row=row, column=3).value = "=$B$6/12*0.12"
    
    # Tax Deduction (Assumed 10% of new Gross Income)
    ws.cell(row=row, column=4).value = "=$B$6*0.10/12"
    
    # Variable Bonus in March only
    if row == 25:  # March
        ws.cell(row=row, column=5).value = "=$B$3"
    else:
        ws.cell(row=row, column=5).value = 0
    
    # Leave Encashment in January (next year)
    if row == 34:  # January
        ws.cell(row=row, column=6).value = "=$B$7*$B$8"
    else:
        ws.cell(row=row, column=6).value = 0
    
    # Spot Cash Award (carry forward same as current year)
    ws.cell(row=row, column=7).value = "=IF(ISBLANK($B$4),0,$B$4)"
    
    # Total Take-Home
    ws.cell(row=row, column=8).value = "=B{row}-C{row}-D{row}+E{row}+F{row}+G{row}".format(row=row)

# Save the final workbook
file_path = 'salary_takeaway_with_increment_leave_encashment.xlsx'
wb.save(file_path)
